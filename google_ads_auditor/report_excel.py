"""
Excel Report Generator - Creates detailed XLSX audit report with multiple tabs.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime


# Color scheme
COLORS = {
    "dark_blue": "1B365D",
    "blue": "2E75B6",
    "light_blue": "D6E4F0",
    "green": "70AD47",
    "light_green": "E2EFDA",
    "yellow": "FFC000",
    "light_yellow": "FFF2CC",
    "orange": "ED7D31",
    "red": "FF0000",
    "light_red": "FCE4EC",
    "white": "FFFFFF",
    "light_gray": "F2F2F2",
    "gray": "808080",
    "dark_gray": "404040",
}

GRADE_COLORS = {
    "A": "70AD47",
    "B": "92D050",
    "C": "FFC000",
    "D": "ED7D31",
    "F": "FF0000",
}

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header_row(ws, row, max_col, fill_color=None):
    fill = PatternFill("solid", fgColor=fill_color or COLORS["dark_blue"])
    font = Font(bold=True, color=COLORS["white"], size=11)
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_data_row(ws, row, max_col, alt=False):
    fill = PatternFill("solid", fgColor=COLORS["light_gray"]) if alt else PatternFill("solid", fgColor=COLORS["white"])
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _auto_width(ws, min_width=10, max_width=50):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def generate_excel_report(audit_results, search_term_analysis, account_info, config, output_path):
    wb = Workbook()

    _create_dashboard_sheet(wb, audit_results, account_info)
    _create_category_breakdown_sheet(wb, audit_results)
    _create_failed_checks_sheet(wb, audit_results)
    _create_quick_wins_sheet(wb, audit_results)
    _create_search_terms_sheet(wb, search_term_analysis)
    _create_negative_keywords_sheet(wb, search_term_analysis)
    _create_keyword_expansion_sheet(wb, search_term_analysis)
    _create_all_checks_sheet(wb, audit_results)

    wb.save(output_path)
    return output_path


def _create_dashboard_sheet(wb, results, account_info):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = COLORS["dark_blue"]

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "GOOGLE ADS AUDIT REPORT"
    ws["A1"].font = Font(bold=True, size=20, color=COLORS["dark_blue"])
    ws["A1"].alignment = Alignment(horizontal="center")

    # Account info
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Account: {account_info.get('name', 'N/A')} | ID: {account_info.get('id', 'N/A')} | Generated: {datetime.now().strftime('%B %d, %Y %H:%M')}"
    ws["A2"].font = Font(size=11, color=COLORS["gray"])
    ws["A2"].alignment = Alignment(horizontal="center")

    # Overall Score Box
    row = 4
    ws.merge_cells(f"B{row}:D{row+3}")
    score_cell = ws[f"B{row}"]
    score_cell.value = results["overall_score"]
    score_cell.font = Font(bold=True, size=48, color=COLORS["white"])
    score_cell.fill = PatternFill("solid", fgColor=GRADE_COLORS.get(results["grade"], COLORS["gray"]))
    score_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"E{row}:G{row+1}")
    ws[f"E{row}"] = f"Grade: {results['grade']}"
    ws[f"E{row}"].font = Font(bold=True, size=28, color=GRADE_COLORS.get(results["grade"], COLORS["dark_gray"]))
    ws[f"E{row}"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"E{row+2}:G{row+3}")
    ws[f"E{row+2}"] = results["grade_label"]
    ws[f"E{row+2}"].font = Font(size=16, color=COLORS["dark_gray"])
    ws[f"E{row+2}"].alignment = Alignment(horizontal="center", vertical="center")

    # Summary Stats
    row = 9
    stats = [
        ("Total Checks", results["total_checks"]),
        ("Passed", results["passed_checks"]),
        ("Failed", results["failed_checks_count"]),
        ("Quick Wins", len(results["quick_wins"])),
    ]
    for i, (label, value) in enumerate(stats):
        col = 2 + i * 2
        ws.cell(row=row, column=col, value=label).font = Font(bold=True, size=11, color=COLORS["gray"])
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        val_cell = ws.cell(row=row + 1, column=col, value=value)
        val_cell.font = Font(bold=True, size=22, color=COLORS["dark_blue"])
        val_cell.alignment = Alignment(horizontal="center")

    # Category Scores
    row = 12
    ws.cell(row=row, column=1, value="CATEGORY BREAKDOWN").font = Font(bold=True, size=14, color=COLORS["dark_blue"])
    row += 1

    headers = ["Category", "Weight", "Checks", "Passed", "Failed", "Score", "Grade"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1, value=h)
    _style_header_row(ws, row, len(headers))

    cat_names = {
        "conversion_tracking": "Conversion Tracking",
        "wasted_spend": "Wasted Spend & Negatives",
        "account_structure": "Account Structure",
        "keywords_qs": "Keywords & Quality Score",
        "ads_assets": "Ads & Assets",
        "settings_targeting": "Settings & Targeting",
    }

    for i, (cat, details) in enumerate(results["category_details"].items()):
        r = row + 1 + i
        score = details["score"]
        from .auditor import grade_from_score
        g, _ = grade_from_score(score)

        ws.cell(row=r, column=1, value=cat_names.get(cat, cat))
        ws.cell(row=r, column=2, value=f"{details['weight']*100:.0f}%")
        ws.cell(row=r, column=3, value=details["checks"])
        ws.cell(row=r, column=4, value=details["passed"])
        ws.cell(row=r, column=5, value=details.get("failed", 0))
        score_cell = ws.cell(row=r, column=6, value=f"{score:.1f}")
        grade_cell = ws.cell(row=r, column=7, value=g)

        _style_data_row(ws, r, len(headers), alt=(i % 2 == 1))
        grade_cell.fill = PatternFill("solid", fgColor=GRADE_COLORS.get(g, COLORS["gray"]))
        grade_cell.font = Font(bold=True, color=COLORS["white"])
        grade_cell.alignment = Alignment(horizontal="center")

    # Add chart
    chart_row = row + 1
    chart = BarChart()
    chart.type = "col"
    chart.title = "Category Scores"
    chart.y_axis.title = "Score"
    chart.y_axis.scaling.max = 100
    chart.x_axis.title = "Category"
    chart.style = 10

    data = Reference(ws, min_col=6, min_row=row, max_row=row + len(results["category_details"]))
    cats = Reference(ws, min_col=1, min_row=row + 1, max_row=row + len(results["category_details"]))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 20
    chart.height = 12
    ws.add_chart(chart, f"A{row + len(results['category_details']) + 2}")

    _auto_width(ws)


def _create_category_breakdown_sheet(wb, results):
    ws = wb.create_sheet("Category Details")
    ws.sheet_properties.tabColor = COLORS["blue"]

    cat_names = {
        "conversion_tracking": "Conversion Tracking",
        "wasted_spend": "Wasted Spend & Negatives",
        "account_structure": "Account Structure",
        "keywords_qs": "Keywords & Quality Score",
        "ads_assets": "Ads & Assets",
        "settings_targeting": "Settings & Targeting",
    }

    row = 1
    for cat, cat_name in cat_names.items():
        cat_checks = [c for c in results["all_checks"] if c["category"] == cat]
        if not cat_checks:
            continue

        ws.merge_cells(f"A{row}:G{row}")
        ws.cell(row=row, column=1, value=cat_name.upper()).font = Font(bold=True, size=14, color=COLORS["dark_blue"])
        row += 1

        headers = ["Check ID", "Check Name", "Severity", "Status", "Details", "Recommendation", "Fix Time"]
        for i, h in enumerate(headers):
            ws.cell(row=row, column=i + 1, value=h)
        _style_header_row(ws, row, len(headers))
        row += 1

        for i, check in enumerate(cat_checks):
            ws.cell(row=row, column=1, value=check["check_id"])
            ws.cell(row=row, column=2, value=check["name"])
            ws.cell(row=row, column=3, value=check["severity"].upper())
            status = "PASS" if check["passed"] else "FAIL"
            status_cell = ws.cell(row=row, column=4, value=status)
            ws.cell(row=row, column=5, value=check["details"])
            ws.cell(row=row, column=6, value=check["recommendation"])
            ws.cell(row=row, column=7, value=f"{check['fix_time_minutes']} min" if check["fix_time_minutes"] else "")

            _style_data_row(ws, row, len(headers), alt=(i % 2 == 1))

            if check["passed"]:
                status_cell.fill = PatternFill("solid", fgColor=COLORS["light_green"])
                status_cell.font = Font(bold=True, color=COLORS["green"])
            else:
                status_cell.fill = PatternFill("solid", fgColor=COLORS["light_red"])
                status_cell.font = Font(bold=True, color=COLORS["red"])
            status_cell.alignment = Alignment(horizontal="center")

            row += 1

        row += 1  # Gap between categories

    _auto_width(ws)


def _create_failed_checks_sheet(wb, results):
    ws = wb.create_sheet("Failed Checks")
    ws.sheet_properties.tabColor = COLORS["red"]

    ws.merge_cells("A1:G1")
    ws["A1"] = f"FAILED CHECKS ({results['failed_checks_count']})"
    ws["A1"].font = Font(bold=True, size=14, color=COLORS["red"])

    headers = ["Priority", "Check ID", "Check Name", "Severity", "Category", "Details", "Recommendation"]
    row = 3
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1, value=h)
    _style_header_row(ws, row, len(headers), COLORS["red"])

    for i, check in enumerate(results["failed_checks"]):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=check["check_id"])
        ws.cell(row=r, column=3, value=check["name"])
        sev_cell = ws.cell(row=r, column=4, value=check["severity"].upper())
        ws.cell(row=r, column=5, value=check["category"].replace("_", " ").title())
        ws.cell(row=r, column=6, value=check["details"])
        ws.cell(row=r, column=7, value=check["recommendation"])

        _style_data_row(ws, r, len(headers), alt=(i % 2 == 1))

        if check["severity"] == "critical":
            sev_cell.fill = PatternFill("solid", fgColor=COLORS["light_red"])
            sev_cell.font = Font(bold=True, color=COLORS["red"])
        elif check["severity"] == "high":
            sev_cell.fill = PatternFill("solid", fgColor=COLORS["light_yellow"])
            sev_cell.font = Font(bold=True, color=COLORS["orange"])

    _auto_width(ws)


def _create_quick_wins_sheet(wb, results):
    ws = wb.create_sheet("Quick Wins")
    ws.sheet_properties.tabColor = COLORS["green"]

    ws.merge_cells("A1:F1")
    ws["A1"] = f"QUICK WINS ({len(results['quick_wins'])} items, fix in <15 min each)"
    ws["A1"].font = Font(bold=True, size=14, color=COLORS["green"])

    headers = ["#", "Check ID", "Issue", "Severity", "Impact", "Est. Fix Time"]
    row = 3
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1, value=h)
    _style_header_row(ws, row, len(headers), COLORS["green"])

    for i, check in enumerate(results["quick_wins"]):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=check["check_id"])
        ws.cell(row=r, column=3, value=f"{check['name']}: {check['recommendation']}")
        ws.cell(row=r, column=4, value=check["severity"].upper())
        ws.cell(row=r, column=5, value=check.get("impact", "Performance improvement"))
        ws.cell(row=r, column=6, value=f"{check['fix_time_minutes']} min")
        _style_data_row(ws, r, len(headers), alt=(i % 2 == 1))

    _auto_width(ws)


def _create_search_terms_sheet(wb, analysis):
    ws = wb.create_sheet("Search Term Report")
    ws.sheet_properties.tabColor = COLORS["orange"]

    ws.merge_cells("A1:H1")
    ws["A1"] = "SEARCH TERM ANALYSIS"
    ws["A1"].font = Font(bold=True, size=14, color=COLORS["orange"])

    # Summary stats
    row = 3
    stats = [
        ("Total Search Terms", analysis["total_terms"]),
        ("Total Cost", f"${analysis['total_cost']:,.2f}"),
        ("Total Conversions", f"{analysis['total_conversions']:.0f}"),
        ("Negative Candidates", analysis["total_negative_candidates"]),
        ("Wasted Spend", f"${analysis['total_wasted_spend']:,.2f}"),
        ("Expansion Candidates", analysis["total_expansion_candidates"]),
    ]
    for i, (label, value) in enumerate(stats):
        ws.cell(row=row, column=1 + (i % 3) * 3, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2 + (i % 3) * 3, value=value).font = Font(size=12, color=COLORS["dark_blue"])
        if i == 2:
            row += 1

    # Category breakdown
    row += 2
    ws.cell(row=row, column=1, value="CATEGORY BREAKDOWN").font = Font(bold=True, size=12, color=COLORS["dark_blue"])
    row += 1

    if analysis["category_summary"]:
        headers = ["Category", "Count", "Total Cost"]
        for i, h in enumerate(headers):
            ws.cell(row=row, column=i + 1, value=h)
        _style_header_row(ws, row, len(headers), COLORS["orange"])
        row += 1

        for cat, data in analysis["category_summary"].items():
            ws.cell(row=row, column=1, value=cat.replace("_", " ").title())
            ws.cell(row=row, column=2, value=data["count"])
            ws.cell(row=row, column=3, value=f"${data['total_cost']:,.2f}")
            _style_data_row(ws, row, len(headers))
            row += 1

    # Top wasted spend terms
    row += 1
    ws.cell(row=row, column=1, value="TOP WASTED SPEND SEARCH TERMS").font = Font(bold=True, size=12, color=COLORS["red"])
    row += 1

    headers = ["Search Term", "Campaign", "Clicks", "Cost", "Conversions", "CTR%", "Avg CPC"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1, value=h)
    _style_header_row(ws, row, len(headers), COLORS["red"])
    row += 1

    for i, term in enumerate(analysis["wasted_spend_terms"][:30]):
        ws.cell(row=row, column=1, value=term["search_term"])
        ws.cell(row=row, column=2, value=term["campaign_name"])
        ws.cell(row=row, column=3, value=term["clicks"])
        ws.cell(row=row, column=4, value=f"${term['cost']:.2f}")
        ws.cell(row=row, column=5, value=term["conversions"])
        ws.cell(row=row, column=6, value=f"{term['ctr']:.2f}%")
        ws.cell(row=row, column=7, value=f"${term['avg_cpc']:.2f}" if term["avg_cpc"] else "N/A")
        _style_data_row(ws, row, len(headers), alt=(i % 2 == 1))
        row += 1

    _auto_width(ws)


def _create_negative_keywords_sheet(wb, analysis):
    ws = wb.create_sheet("Negative Keyword Suggestions")
    ws.sheet_properties.tabColor = COLORS["red"]

    ws.merge_cells("A1:H1")
    ws["A1"] = f"NEGATIVE KEYWORD SUGGESTIONS ({analysis['total_negative_candidates']} candidates)"
    ws["A1"].font = Font(bold=True, size=14, color=COLORS["red"])

    ws.merge_cells("A2:H2")
    ws["A2"] = "REVIEW BEFORE ADDING - These are suggestions based on performance data and pattern matching"
    ws["A2"].font = Font(italic=True, color=COLORS["gray"])

    headers = ["Priority", "Search Term", "Suggested Negative", "Match Type", "Campaign",
               "Cost", "Clicks", "Reason"]
    row = 4
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1, value=h)
    _style_header_row(ws, row, len(headers))

    from .search_term_analyzer import generate_negative_keyword_suggestions
    suggestions = generate_negative_keyword_suggestions(analysis)

    for i, sug in enumerate(suggestions):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=sug["triggered_by"])
        ws.cell(row=r, column=3, value=sug["negative_keyword"])
        ws.cell(row=r, column=4, value=sug["suggested_match_type"])
        ws.cell(row=r, column=5, value="")
        ws.cell(row=r, column=6, value=f"${sug['cost_impact']:.2f}")
        ws.cell(row=r, column=7, value="")
        ws.cell(row=r, column=8, value="; ".join(sug["reasons"]))
        _style_data_row(ws, r, len(headers), alt=(i % 2 == 1))

    _auto_width(ws)


def _create_keyword_expansion_sheet(wb, analysis):
    ws = wb.create_sheet("Keyword Expansion")
    ws.sheet_properties.tabColor = COLORS["green"]

    ws.merge_cells("A1:G1")
    ws["A1"] = f"KEYWORD EXPANSION CANDIDATES ({analysis['total_expansion_candidates']})"
    ws["A1"].font = Font(bold=True, size=14, color=COLORS["green"])

    ws.merge_cells("A2:G2")
    ws["A2"] = "Search terms converting that are not yet added as keywords"
    ws["A2"].font = Font(italic=True, color=COLORS["gray"])

    headers = ["Search Term", "Campaign", "Impressions", "Clicks", "Cost", "Conversions", "CPA"]
    row = 4
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1, value=h)
    _style_header_row(ws, row, len(headers), COLORS["green"])

    for i, cand in enumerate(analysis["keyword_expansion_candidates"][:30]):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=cand["search_term"])
        ws.cell(row=r, column=2, value=cand["campaign_name"])
        ws.cell(row=r, column=3, value=cand["impressions"])
        ws.cell(row=r, column=4, value=cand["clicks"])
        ws.cell(row=r, column=5, value=f"${cand['cost']:.2f}")
        ws.cell(row=r, column=6, value=cand["conversions"])
        ws.cell(row=r, column=7, value=f"${cand['cpa']:.2f}" if cand["cpa"] else "N/A")
        _style_data_row(ws, r, len(headers), alt=(i % 2 == 1))

    _auto_width(ws)


def _create_all_checks_sheet(wb, results):
    ws = wb.create_sheet("All Checks (74)")
    ws.sheet_properties.tabColor = COLORS["gray"]

    headers = ["Check ID", "Check Name", "Category", "Severity", "Status", "Details"]
    row = 1
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1, value=h)
    _style_header_row(ws, row, len(headers))

    for i, check in enumerate(results["all_checks"]):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=check["check_id"])
        ws.cell(row=r, column=2, value=check["name"])
        ws.cell(row=r, column=3, value=check["category"].replace("_", " ").title())
        ws.cell(row=r, column=4, value=check["severity"].upper())
        status = "PASS" if check["passed"] else "FAIL"
        status_cell = ws.cell(row=r, column=5, value=status)
        ws.cell(row=r, column=6, value=check["details"])

        _style_data_row(ws, r, len(headers), alt=(i % 2 == 1))
        if check["passed"]:
            status_cell.font = Font(bold=True, color=COLORS["green"])
        else:
            status_cell.font = Font(bold=True, color=COLORS["red"])

    _auto_width(ws)
